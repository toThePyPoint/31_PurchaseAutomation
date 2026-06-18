import pandas as pd
from openpyxl import load_workbook
import openpyxl
import datetime
import os
import glob
from pathlib import Path
import time
import logging


def get_supplier_sap_numbers(filepath, sheet_name=None):
    """
    Returns a list of SAP numbers from an Excel file, which are found in column A
    between the row containing 'SAP' and the row containing 'Consumption'.

    :param filepath: Path to the Excel file (.xlsx)
    :param sheet_name: Name of the worksheet (if None, the active sheet is used)
    :return: List of SAP numbers (values from column A)
    """
    # Open the Excel file
    print("Filepath: ", filepath)
    print("sheet_name: ", sheet_name)
    wb = openpyxl.load_workbook(filepath, keep_vba=True)
    if sheet_name:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # Search for the headers in column A
    row_sap = None
    row_consumption = None
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value
        if cell_value is not None:
            value_str = str(cell_value).strip().upper()
            if value_str == 'SAP' and row_sap is None:
                row_sap = row
            elif value_str == 'CONSUMPTION' and row_consumption is None:
                row_consumption = row
        if row_sap is not None and row_consumption is not None:
            break

    if row_sap is None or row_consumption is None:
        return []  # Or raise an Exception if you prefer

    # Get numbers between the found headers (exclusive)
    sap_numbers = []
    for row in range(row_sap + 1, row_consumption):
        value = sheet.cell(row=row, column=1).value
        if value is not None:
            sap_numbers.append(str(value))
    return sap_numbers

def get_zek103_data(zek103_data, plant, sap_numbers):
    zek103_data = zek103_data[zek103_data['Werk'] == plant]
    zek103_data = zek103_data[zek103_data['Mat'].isin(sap_numbers)]

    # Implement the logic which ensures operating on appropriate set of data (date of delivery and order quantity)
    zek103_data = zek103_data.copy()
    confirmed_delivery_date_present = (
        zek103_data['Best. Liefdat.'].notna()
        & zek103_data['Best. Liefdat.'].astype(str).str.strip().ne('')
    )
    confirmed_quantity_lower = zek103_data['Off. Mg'] > zek103_data['Bestät. Menge']

    zek103_data['_output_lieferdatum'] = zek103_data['Lieferdatum']
    zek103_data['_output_off_mg'] = zek103_data['Off. Mg']

    zek103_data.loc[confirmed_quantity_lower, '_output_off_mg'] = zek103_data.loc[
        confirmed_quantity_lower, 'Bestät. Menge'
    ]
    zek103_data.loc[confirmed_delivery_date_present, '_output_lieferdatum'] = zek103_data.loc[
        confirmed_delivery_date_present, 'Best. Liefdat.'
    ]
    zek103_data.loc[confirmed_delivery_date_present, '_output_off_mg'] = zek103_data.loc[
        confirmed_delivery_date_present, 'Bestät. Menge'
    ]

    zek103_data_grouped = (
        zek103_data
        .groupby(['_output_lieferdatum', 'Mat'], as_index=False)['_output_off_mg']
        .sum()
        .rename(columns={
            '_output_lieferdatum': 'Lieferdatum',
            '_output_off_mg': 'Off. Mg',
        })
    )

    zek103_data_grouped['Lieferdatum'] = pd.to_datetime(zek103_data_grouped['Lieferdatum'])  # opcjonalna konwersja na datę
    zek103_data_grouped['delayed'] = zek103_data_grouped['Lieferdatum'] < pd.Timestamp('today').normalize()

    print("ZEK103_data: ", zek103_data_grouped)
    return zek103_data_grouped

def update_excel_with_quantities(filepath, df, header_upper_bound, header_lower_bound, sheet_name, is_order_data=False):
    """
    Updates an Excel file with quantities from a DataFrame based on matching SAP numbers
    and Lieferdatum (dates).

    :param filepath: Path to the Excel file
    :param df: DataFrame with columns ['Lieferdatum', 'Mat', 'Best-Mg', 'delayed']
    """
    # Load the workbook
    wb = openpyxl.load_workbook(filepath, keep_vba=True)
    # sheet = wb.active  # Operate on the active sheet
    sheet = wb[sheet_name]

    # Find the row range for SAP numbers (between "SAP" and "Consumption")
    sap_start_row = None
    sap_end_row = None
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=1).value  # Column A
        if cell_value is not None:
            value_str = str(cell_value).strip().upper()
            if value_str == header_upper_bound and sap_start_row is None:
                sap_start_row = row + 1  # Start after "SAP"
            elif value_str == header_lower_bound and sap_end_row is None:
                sap_end_row = row - 1  # End before "Consumption"

    if sap_start_row is None or sap_end_row is None:
        raise ValueError(f"Unable to find {header_upper_bound} and {header_lower_bound} headers in column A.")

    # Read SAP numbers from the specified range in column A
    sap_numbers = {}
    for row in range(sap_start_row, sap_end_row + 1):
        mat_number = sheet.cell(row=row, column=1).value
        if mat_number is not None:
            sap_numbers[str(mat_number)] = row

    # Read dates from the first row (headers starting from column M)
    date_columns = {'delayed': 12}  # Column with delayed orders

    for col in range(13, sheet.max_column + 1):  # Column index starts at M (13th column)
        date_value = sheet.cell(row=1, column=col).value
        if isinstance(date_value, pd.Timestamp) or isinstance(date_value, datetime.date):
            date_value = pd.Timestamp(date_value)  # Ensure it is a pandas Timestamp
        if date_value:
            date_columns[date_value.date()] = col

    # Clear data in rows between 'SAP' and 'Consumption', from column L to column L+300 (12th to 312th column)
    for row in range(sap_start_row, sap_end_row + 1):
        for col in range(12, 312):  # Column L (12th column) to 312th column
            sheet.cell(row=row, column=col).value = None

    # Iterate over the DataFrame rows
    for _, row in df.iterrows():
        sap_row = None
        date_col = None

        lieferdatum = row['Lieferdatum']
        mat_number = row['Mat']
        quantity = row['Off. Mg']
        delayed = row['delayed']

        # Ensure lieferdatum is a date (convert if necessary)
        if isinstance(lieferdatum, str):
            lieferdatum = pd.to_datetime(lieferdatum).date()
        elif isinstance(lieferdatum, pd.Timestamp):
            lieferdatum = lieferdatum.date()

        # Match SAP number and date to find the correct cell
        if mat_number in sap_numbers and lieferdatum in date_columns:
            sap_row = sap_numbers[mat_number]
            date_col = date_columns[lieferdatum]
        if delayed and is_order_data:
            sap_row = sap_numbers[mat_number]
            date_col = date_columns['delayed']

        if sap_row is not None and date_col is not None:
            # Write quantity to the matched cell
            sheet.cell(row=sap_row, column=date_col).value = quantity

    # Save the updated workbook
    wb.save(filepath)
    wb.close()
    print(f"Excel file {os.path.basename(filepath)} updated successfully.")

def list_excel_files(directory):
    """
    Returns a list of full paths to all Excel files (.xlsx and .xlsm) in a specified directory.

    :param directory: Path to the directory to search for Excel files
    :return: List of full paths to Excel files
    """

    # Use glob to find .xlsx and .xlsm files in the directory
    xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))
    xlsm_files = glob.glob(os.path.join(directory, '*.xlsm'))

    # Combine and filter out temporary Excel files
    excel_files = [
        f for f in (xlsx_files + xlsm_files)
        if not os.path.basename(f).startswith('~$')
    ]

    return excel_files

def get_mb52_data(file_path, dtypes):

    df = pd.read_excel(file_path, dtype=dtypes)
    return df

def filter_mb52_data(df, sap_numbers, plant='2101', storage_locs='0007', col_name='Frei verwendbar'):
    # Ensure storage locks is a tuple (convert if necessary)
    if isinstance(storage_locs, str):
        storage_locs = storage_locs,

    df = df[(df['Lagerort'].isin(storage_locs)) & (df['Werk'] == plant) & (df['Material'].isin(sap_numbers))]

    df_grouped = df.groupby(['Material'], as_index=False)[col_name].sum()

    return df_grouped

def update_excel_with_dataframe(file_path, dataframe, sap_column, frei_column, header_start, header_end, sheet_name, col_name='Frei verwendbar'):
    """
    Uzupełnia plik Excel danymi z DataFrame w określonym zakresie wierszy między nagłówkami.

    :param file_path: Ścieżka do pliku Excel
    :param dataframe: DataFrame zawierający dane (z kolumnami 'Material' i col_name)
    :param sap_column: Nazwa kolumny z numerami SAP w Excelu (np. 'A')
    :param frei_column: Nazwa kolumny w Excelu, do której mają być wpisywane dane (np. 'K')
    :param header_start: Nagłówek wskazujący początek zakresu (np. 'Stock')
    :param header_end: Nagłówek wskazujący koniec zakresu (np. 'S.C')
    """
    # Załadowanie istniejącego pliku Excel
    workbook = load_workbook(filename=file_path, keep_vba=True)
    # sheet = workbook.active  # Zakładamy, że pracujemy na aktywnym arkuszu
    sheet = workbook[sheet_name]

    # Znalezienie zakresów wierszy na podstawie nagłówków
    start_row = None
    end_row = None

    for row in sheet.iter_rows():
        cell_a_value = row[0].value  # Wartość w kolumnie A (zakładamy, że A to sap_column)
        if cell_a_value == header_start:
            start_row = row[0].row + 1  # Zakres zaczyna się od wiersza poniżej nagłówka
        elif cell_a_value == header_end:
            end_row = row[0].row - 1  # Kończy się przed wierszem z nagłówkiem
            break

    if start_row is None or end_row is None:
        print("Nie znaleziono odpowiednich nagłówków w pliku Excel.")
        return

    # Przekształcenie DataFrame na dictionary dla szybkiego wyszukiwania
    data_mapping = dict(zip(dataframe['Material'], dataframe[col_name]))

    # Iteracja po podanym zakresie wierszy w Excelu
    for row in range(start_row, end_row + 1):
        sap_value = str(sheet[f'{sap_column}{row}'].value)  # Pobierz wartość z kolumny SAP
        if sap_value in data_mapping:  # Jeśli wartość SAP występuje w DataFrame
            sheet[f'{frei_column}{row}'] = data_mapping[sap_value]  # Wpisz wartość z kolumny col_name
        else:
            sheet[f'{frei_column}{row}'] = None

    # Zapisanie pliku Excel
    workbook.save(file_path)
    workbook.close()
    print(f"Plik {file_path} został zaktualizowany.")

def get_mb51_consumption_data(z_mat_file_path, dtypes, col_names):
    df = pd.read_excel(z_mat_file_path, dtype=dtypes)
    df = df.rename(columns=col_names)
    df = df.groupby(['Material', 'date', 'plant'], as_index=False)['quantity'].sum()
    df['date'] = pd.to_datetime(df['date']).dt.date  # Konwersja na `datetime.date`

    return df

def get_past_workdays(start_date, num_days, mode="last", country_holidays=None):
    """
    mode: "last" --> days before the start date, "next" --> days after the start date
    Zwraca listę dat num_days dni roboczych wstecz od start_date.
    Pomija weekendy i opcjonalnie święta.
    """
    workdays = []
    current_date = start_date

    while len(workdays) < num_days:
        if mode == 'last':
            current_date -= datetime.timedelta(days=1)
        elif mode == "next":
            current_date += datetime.timedelta(days=1)
        # Sprawdź, czy to dzień roboczy
        if current_date.weekday() < 5:  # poniedziałek=0, piątek=4
            if country_holidays is None or current_date not in country_holidays:
                workdays.append(current_date)

    return workdays

def filter_z_mat_consumption_data(consumption_df, mat_list, days_range, plant):
    consumption_df_filtered = consumption_df[(consumption_df['date'].isin(days_range)) & (consumption_df['Material'].isin(mat_list)) & (consumption_df['plant'] == plant)]
    consumption_df_grouped = consumption_df_filtered.groupby('Material', as_index=False)['quantity'].sum()

    return consumption_df_grouped

def extract_unique_mb51_files(data_dict: dict) -> dict:
    result = {}

    for file_list in data_dict.values():
        if not file_list:
            continue

        key = file_list[0]

        if key not in result:
            result[key] = []

        for item in file_list:
            if item not in result[key]:
                result[key].append(item)

    return result

def load_grouped_mb51_dataframes(
    grouped_files: dict,
    base_path: str,
    extension: str = ".xlsx",
    dtypes_mb51: dict | None = None,
    col_names_mb51: dict | None = None,
    ignore_missing: bool = False,
) -> dict:
    """
    Load and combine files into DataFrames grouped by key.

    :param col_names_mb51:
    :param dtypes_mb51:
    :param dtypes:
    :param grouped_files: dict, e.g. {"CONS_246": ["CONS_246", "CONS_246_2"]}
    :param base_path: directory path containing the files
    :param extension: file extension (".xlsx", ".csv", etc.)
    :param ignore_missing: if True, skip missing files instead of raising error
    :return: dict[str, pd.DataFrame]
    """

    base_path = Path(base_path)
    result = {}

    for group_key, file_list in grouped_files.items():
        print(f"Processing {group_key}")
        dataframes = []

        for file_name in file_list:
            file_path = base_path / f"{file_name}{extension}"

            if not file_path.exists():
                if ignore_missing:
                    continue
                else:
                    raise FileNotFoundError(f"File not found: {file_path}")

            # Select appropriate loader
            if extension.upper() == ".XLSX":
                df = get_mb51_consumption_data(file_path, dtypes_mb51, col_names_mb51)
            else:
                raise ValueError(f"Unsupported file extension: {extension}")

            # Add metadata columns (useful for debugging / traceability)
            df["source_file"] = file_name
            df["group_key"] = group_key

            dataframes.append(df)

        # Combine all DataFrames for the group
        if dataframes:
            result[group_key] = pd.concat(dataframes, ignore_index=True)
        else:
            result[group_key] = pd.DataFrame()

    return result

def retrieve_supplier_name(file_path: str) -> str:
    """
    Extract supplier name from full file path.

    Example:
    'P:\\...\\PurchAutomation_2101_ABC_COLORE.xlsm' -> 'ABC_COLORE'
    """

    file_name = Path(file_path).name  # extract file name

    start = file_name.find("_")
    end = file_name.rfind(".")

    if start == -1 or end == -1 or start >= end:
        raise ValueError(f"Invalid file name format: {file_path}")

    return file_name[start + 6:end]

def retrieve_plant(file_path: str) -> str:
    """
    Extract plant code from full file path.

    Example:
    'P:\\...\\PurchAutomation_2101_ABC_COLORE.xlsm' -> 'ABC_COLORE'
    """

    file_name = Path(file_path).name  # extract file name

    start = file_name.find("_")
    end = start + 5

    if start == -1 or end == -1 or start >= end:
        raise ValueError(f"Invalid file name format: {file_path}")

    return file_name[start + 1:end]


def retry(
    func,
    retries=5,
    delay=2,
    exceptions=(PermissionError, OSError),
    *args,
    **kwargs
):
    """
    Retry wrapper for unstable file/network operations.
    """

    for attempt in range(1, retries + 1):

        try:
            return func(*args, **kwargs)

        except exceptions as e:

            logging.warning(
                f"{type(e).__name__} in {func.__name__} "
                f"(attempt {attempt}/{retries}): {e}"
            )

            if attempt == retries:
                logging.error(
                    f"Final failure in {func.__name__}",
                    exc_info=True
                )
                raise

            wait_time = delay * attempt

            print(
                f"{type(e).__name__} detected "
                f"for file operation. "
                f"Retry in {wait_time}s..."
            )

            time.sleep(wait_time)
